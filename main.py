import tkinter as tk
from tkinter import filedialog, ttk
from tkinter import font
import pandas as pd
from openpyxl import load_workbook
from construction_analyzer import ConstructionAnalyzer
from design_analyzer import DesignAnalyzer
from excel_saver import ExcelResultSaver
from leader_analyzer import LeaderFrequencyAnalyzer
from loss_analyzer import LossDataAnalyzer
from loss_over_analyzer import LossOverAnalyzer


class ModernButton(tk.Button):
    """自定义圆角按钮"""

    def __init__(self, parent, text, command, bg="#4a90e2", fg="white",
                 hover_bg="#3a80d2", width=12, height=2, **kwargs):
        super().__init__(parent, text=text, command=command,
                         bg=bg, fg=fg, borderwidth=0,
                         width=width, height=height, **kwargs)
        self.hover_bg = hover_bg
        self.original_bg = bg
        self.bind("<Enter>", self.on_enter)
        self.bind("<Leave>", self.on_leave)
        self.config(font=("微软雅黑", 10, "bold"))

    def on_enter(self, e):
        self['background'] = self.hover_bg

    def on_leave(self, e):
        self['background'] = self.original_bg


class MainApp:
    def __init__(self, root):
        self.root = root
        self.root.title("亏损数据分析工具")
        self.root.geometry("900x650")
        self.root.minsize(800, 600)

        # 设置中文字体
        self.setup_fonts()

        # 颜色方案
        self.colors = {
            "primary": "#2c3e50",
            "secondary": "#3498db",
            "accent": "#e74c3c",
            "light": "#ecf0f1",
            "dark": "#34495e",
            "success": "#2ecc71",
            "error": "#e74c3c",
            "info": "#3498db"
        }

        self.original_columns = []  # 原始表头列名
        self.raw_data = None  # 原始数据DataFrame
        self.excel_saver = ExcelResultSaver()
        self.analyzers = []  # 分析器实例列表

        # 分析器配置
        self.analyzers_config = [
            {
                "class": LeaderFrequencyAnalyzer,
                "sheet_name": "附表1  亏损3个项目项目负责人",
                "analyze_kwargs": {"min_count": 3}
            },
            {
                "class": DesignAnalyzer,
                "sheet_name": "附表2  亏损大于合同",
                "analyze_kwargs": {}
            },
            {
                "class": ConstructionAnalyzer,
                "sheet_name": "附表3  施工项目亏损金额占合同金额30%",
                "analyze_kwargs": {}
            },
            {
                "class": LossOverAnalyzer,
                "sheet_name": "附表4  亏损大于1000万",
                "analyze_kwargs": {"threshold": 1000}
            },
            {
                "class": LossDataAnalyzer,
                "sheet_name": "附表5  成本费用异常情况",
                "analyze_kwargs": {}
            }
        ]

        self.create_ui()
        self.add_animation()

    def setup_fonts(self):
        """设置支持中文的字体"""
        default_font = font.nametofont("TkDefaultFont")
        default_font.configure(family="微软雅黑", size=10)
        self.root.option_add("*Font", default_font)

    def create_ui(self):
        """创建现代化界面组件"""
        # 设置背景渐变
        self.root.configure(bg=self.colors["light"])

        # 创建标题区域
        header_frame = tk.Frame(self.root, bg=self.colors["primary"], height=80)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)

        title_label = tk.Label(
            header_frame,
            text="亏损数据分析工具",
            bg=self.colors["primary"],
            fg="white",
            font=("微软雅黑", 18, "bold")
        )
        title_label.pack(pady=20)

        # 内容区域
        content_frame = tk.Frame(self.root, bg=self.colors["light"])
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # 按钮区域 - 使用卡片式设计
        button_card = tk.Frame(
            content_frame,
            bg="white",
            relief=tk.RAISED,
            bd=1,
            highlightbackground=self.colors["secondary"],
            highlightthickness=1,
            padx=20,
            pady=15
        )
        button_card.pack(fill=tk.X, pady=(0, 15))

        # 按钮容器
        btn_frame = tk.Frame(button_card, bg="white")
        btn_frame.pack()

        # 上传按钮
        self.upload_btn = ModernButton(
            btn_frame,
            text="上传Excel",
            command=self.upload_excel,
            bg=self.colors["secondary"]
        )
        self.upload_btn.grid(row=0, column=0, padx=10)

        # 分析按钮
        self.analyze_btn = ModernButton(
            btn_frame,
            text="执行分析",
            command=self.run_analysis,
            bg=self.colors["primary"]
        )
        self.analyze_btn.grid(row=0, column=1, padx=10)

        # 保存按钮
        self.save_btn = ModernButton(
            btn_frame,
            text="保存结果",
            command=self.save_results,
            bg=self.colors["accent"]
        )
        self.save_btn.grid(row=0, column=2, padx=10)

        # 日志区域 - 带边框和标题
        log_card = tk.Frame(
            content_frame,
            bg="white",
            relief=tk.RAISED,
            bd=1,
            highlightbackground=self.colors["secondary"],
            highlightthickness=1
        )
        log_card.pack(fill=tk.BOTH, expand=True)

        # 日志标题
        log_title = tk.Label(
            log_card,
            text="操作日志",
            bg=self.colors["dark"],
            fg="white",
            font=("微软雅黑", 10, "bold"),
            padx=10,
            anchor=tk.W
        )
        log_title.pack(fill=tk.X)

        # 日志文本区域
        self.log_frame = tk.Frame(log_card, bg="white")
        self.log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 添加滚动条
        scrollbar = ttk.Scrollbar(self.log_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.log_text = tk.Text(
            self.log_frame,
            height=15,
            width=80,
            yscrollcommand=scrollbar.set,
            wrap=tk.WORD,
            bg="#f9f9f9",
            relief=tk.FLAT,
            padx=5,
            pady=5,
            font=("微软雅黑", 9)
        )
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.log_text.yview)

        # 日志标签样式
        self.log_text.tag_config("info", foreground=self.colors["dark"])
        self.log_text.tag_config("error", foreground=self.colors["error"])
        self.log_text.tag_config("success", foreground=self.colors["success"])
        self.log_text.tag_config("highlight", foreground=self.colors["secondary"], font=("微软雅黑", 9, "bold"))

    def add_animation(self):
        """添加简单的动画效果"""

        def animate_button(button, scale=1.02):
            """按钮点击动画"""
            original_width = button["width"]
            original_height = button["height"]

            button.config(width=int(original_width * scale), height=int(original_height * scale))
            self.root.after(100, lambda: button.config(width=original_width, height=original_height))

        # 为按钮添加点击动画
        for btn in [self.upload_btn, self.analyze_btn, self.save_btn]:
            btn.bind("<ButtonPress-1>", lambda e, b=btn: animate_button(b))

    def get_merged_value(self, sheet, row, col):
        """获取指定行列的单元格值（处理合并单元格）"""
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        return sheet.cell(row=row, column=col).value

    def upload_excel(self):
        """上传Excel并读取数据"""
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel Files", "*.xlsx;*.xls")]
        )
        if not file_path:
            return

        try:
            # 读取表头（3-5行）
            wb = load_workbook(file_path, data_only=True)
            sheet = wb.active
            max_col = sheet.max_column
            self.original_columns = []
            for col_idx in range(1, max_col + 1):
                # 解析三级表头
                main_header = self.get_merged_value(sheet, row=3, col=col_idx)
                sub1_header = self.get_merged_value(sheet, row=4, col=col_idx)
                sub2_header = self.get_merged_value(sheet, row=5, col=col_idx)

                # 去重处理
                parts = []
                seen = set()
                for p in [main_header, sub1_header, sub2_header]:
                    if p is not None:
                        p_str = str(p)
                        if p_str not in seen:
                            seen.add(p_str)
                            parts.append(p_str)

                col_name = "_".join(parts) if parts else f"未知列_{col_idx}"
                self.original_columns.append(col_name)

            self.log(f"三级表头解析完成，共 {len(self.original_columns)} 列", "info")
            self.log(f"示例列名：{self.original_columns[:5]}...", "info")

            # 读取数据行（从第6行开始）
            data = []
            for row in sheet.iter_rows(min_row=6, values_only=True):
                row_data = row[:len(self.original_columns)]
                if row_data and row_data[0] is None:
                    break
                data.append(row_data)
            total_rows = len(data)
            self.log(f"共读取 {total_rows} 行原始数据", "info")
            self.raw_data = pd.DataFrame(data, columns=self.original_columns)

            # 初始化分析器
            self.analyzers = [
                cfg["class"](original_columns=self.original_columns)
                for cfg in self.analyzers_config
            ]
            self.log("分析器初始化完成，可执行分析", "success")

        except Exception as e:
            self.log(f"读取Excel失败：{str(e)}", "error")

    def run_analysis(self):
        """执行所有分析器"""
        if self.raw_data is None or not self.analyzers:
            self.log("请先上传Excel文件！", "error")
            return

        for i, cfg in enumerate(self.analyzers_config):
            analyzer = self.analyzers[i]
            self.log(f"\n===== 开始执行 {analyzer.__class__.__name__} 分析 =====", "highlight")
            success = analyzer.analyze(
                df=self.raw_data,
                parent=self.root, **cfg["analyze_kwargs"]
            )
            if success:
                for log in analyzer.get_logs():
                    self.log(log, "info")
                self.log(f"{analyzer.__class__.__name__} 分析完成", "success")
            else:
                self.log(f"{analyzer.__class__.__name__} 分析失败", "error")

    def save_results(self):
        """保存所有分析结果"""
        if not self.analyzers:
            self.log("请先执行分析！", "error")
            return

        for i, cfg in enumerate(self.analyzers_config):
            analyzer = self.analyzers[i]
            data = analyzer.get_analyzed_data()
            if data is not None and len(data) > 0:
                success = self.excel_saver.save_to_excel(
                    data=data,
                    sheet_base_name=cfg["sheet_name"],
                    parent=self.root
                )
                if success:
                    self.log(f"已保存 {cfg['sheet_name']}", "success")
                else:
                    self.log(f"保存 {cfg['sheet_name']} 失败", "error")
            else:
                self.log(f"{cfg['sheet_name']} 无有效数据，跳过保存", "info")

    def log(self, msg, level="info"):
        """显示日志（支持分级颜色）"""
        timestamp = pd.Timestamp.now().strftime('%H:%M:%S')
        self.log_text.insert(tk.END, f"[{timestamp}] {msg}\n", level)
        self.log_text.see(tk.END)


if __name__ == "__main__":
    root = tk.Tk()
    app = MainApp(root)
    root.mainloop()