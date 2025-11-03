import pandas as pd
import os
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font

class ExcelResultSaver:
    def __init__(self):
        self.file_path = None  # 记录Excel文件路径

    def save_to_excel(self, data, sheet_base_name="数据结果", parent=None):
        """保存数据到Excel（支持多Sheet追加）"""
        if data is None or len(data) == 0:
            return False

        # 确定保存路径
        if self.file_path is None:
            default_filename = f"分析结果_{pd.Timestamp.now().strftime('%Y%m%d%H%M%S')}"
            self.file_path = filedialog.asksaveasfilename(
                title="选择保存路径（后续结果将追加到该文件）",
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile=default_filename,
                parent=parent
            )
            if not self.file_path:  # 用户取消选择
                return False

        # 生成唯一的Sheet名称
        sheet_name = self._get_unique_sheet_name(sheet_base_name)

        # 写入Excel
        try:
            file_exists = os.path.exists(self.file_path)
            mode = "a" if file_exists else "w"

            with pd.ExcelWriter(
                self.file_path,
                engine="openpyxl",
                mode=mode,
                if_sheet_exists="new" if file_exists else None
            ) as writer:
                # 写入数据（预留第1行写说明）
                data.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    startrow=1
                )

                # 添加说明行（加粗）
                worksheet = writer.sheets[sheet_name]
                worksheet["A1"] = (
                    f"数据说明：共{len(data)}行 "
                    f"（生成时间：{pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}）"
                )
                worksheet["A1"].font = Font(bold=True)
            return True

        except PermissionError:
            messagebox.showerror(
                "权限错误",
                "保存失败：文件可能被其他程序占用，请关闭后重试",
                parent=parent
            )
            return False
        except Exception as e:
            messagebox.showerror(
                "错误",
                f"保存失败：{str(e)}",
                parent=parent
            )
            return False

    def _get_unique_sheet_name(self, base_name):
        """生成唯一的Sheet名称（避免重复）"""
        if not os.path.exists(self.file_path):
            return base_name

        # 检查现有Sheet名称
        wb = load_workbook(self.file_path)
        existing_sheets = wb.sheetnames
        wb.close()

        if base_name not in existing_sheets:
            return base_name

        # 重复则添加序号
        suffix = 1
        while f"{base_name}_{suffix}" in existing_sheets:
            suffix += 1
        return f"{base_name}_{suffix}"

    # 新增一个方法用于 DataFrame 存储到 Sheet
    def save_dataframe_to_sheet(self, df: pd.DataFrame, sheet_name: str):
        """将 DataFrame 写入指定的 Sheet 中"""
        try:
            # 处理无效浮动值
            df_cleaned = df.applymap(self.replace_invalid_floats)

            file_exists = os.path.exists(self.file_path)
            mode = "a" if file_exists else "w"

            with pd.ExcelWriter(
                self.file_path,
                engine="openpyxl",
                mode=mode,
                if_sheet_exists="new" if file_exists else None
            ) as writer:
                df_cleaned.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    startrow=1
                )
                worksheet = writer.sheets[sheet_name]
                worksheet["A1"] = f"数据说明：共{len(df_cleaned)}行"
                worksheet["A1"].font = Font(bold=True)
            
            return True
        except Exception as e:
            print(f"写入 {sheet_name} 到 Excel 失败: {e}")
            return False

    def replace_invalid_floats(self, value):
        """替换无效的 float 值（NaN, Infinity）"""
        if isinstance(value, float):
            if pd.isna(value) or value == float('inf') or value == float('-inf'):
                return None  # 或者返回一个替代字符串 "NaN"
        return value
