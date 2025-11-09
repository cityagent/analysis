import json
import os
import pandas as pd
import numpy as np
from fastapi import FastAPI, File, UploadFile, HTTPException
from io import BytesIO
from openpyxl import load_workbook, Workbook
from fastapi.responses import FileResponse, JSONResponse  # <-- 导入 JSONResponse
from fastapi.encoders import jsonable_encoder
import logging
from datetime import datetime
# 导入所需的分析器类
from leader_analyzer import LeaderFrequencyAnalyzer
from construction_analyzer import ConstructionAnalyzer
from design_analyzer import DesignAnalyzer
from loss_over_analyzer import LossOverAnalyzer
from loss_analyzer import LossDataAnalyzer
from excel_saver import ExcelResultSaver

from fastapi.middleware.cors import CORSMiddleware

# 设置日志配置
logging.basicConfig(level=logging.INFO)  # 设置日志级别为 INFO
logger = logging.getLogger(__name__)

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins or specify your allowed origins
    allow_credentials=True,
    allow_methods=["*"],  # Allow all HTTP methods, including OPTIONS
    allow_headers=["*"],  # Allow all headers
)
def convert_datetime_to_string(obj):
    """Recursively convert datetime objects to string in the format of 'YYYY-MM-DD HH:MM:SS'."""
    if isinstance(obj, datetime):
        return obj.strftime('%Y-%m-%d %H:%M:%S')  # Adjust the format as per your requirement
    elif isinstance(obj, dict):
        return {key: convert_datetime_to_string(value) for key, value in obj.items()}
    elif isinstance(obj, list):
        return [convert_datetime_to_string(item) for item in obj]
    return obj
class AnalysisAPI:
    def __init__(self):
        self.original_columns = [] 
        self.raw_data = None  
        self.analyzers = []

        self.analyzers_config = [
            {
                "class": LeaderFrequencyAnalyzer,
                "sheet_name": "亏损3个项目项目负责人",
                "analyze_kwargs": {"min_count": 3}
            },
            {
                "class": DesignAnalyzer,
                "sheet_name": "亏损大于合同",
                "analyze_kwargs": {}
            },
            {
                "class": ConstructionAnalyzer,
                "sheet_name": "施工项目亏损金额占合同金额30%",
                "analyze_kwargs": {}
            },
            {
                "class": LossOverAnalyzer,
                "sheet_name": "亏损大于1000万",
                "analyze_kwargs": {"threshold": 1000}
            },
            {
                "class": LossDataAnalyzer,
                "sheet_name": "成本费用异常情况",
                "analyze_kwargs": {}
            }
        ]

    def upload_excel(self, file: UploadFile):
        """解析上传的Excel文件并读取数据"""
        try:
            content = file.file.read()
            wb = load_workbook(BytesIO(content), data_only=True)
            sheet = wb.active
            max_col = sheet.max_column
            self.original_columns = []
            for col_idx in range(1, max_col + 1):
                main_header = self.get_merged_value(sheet, row=3, col=col_idx)
                sub1_header = self.get_merged_value(sheet, row=4, col=col_idx)
                sub2_header = self.get_merged_value(sheet, row=5, col=col_idx)

                parts = []
                seen = set()
                for p in [main_header, sub1_header, sub2_header]:
                    if p is not None:
                        p_str = str(p)
                        if p_str not in seen:
                            seen.add(p_str)
                            parts.append(p_str)

                col_name = "_".join(parts) if parts else f"未知列_{col_idx}"
                self.original_columns.append(col_name)

            data = []
            for row in sheet.iter_rows(min_row=6, values_only=True):
                row_data = row[:len(self.original_columns)]
                if row_data and row_data[0] is None:
                    break
                data.append(row_data)

            self.raw_data = pd.DataFrame(data, columns=self.original_columns)

            self.analyzers = [
                cfg["class"](original_columns=self.original_columns)
                for cfg in self.analyzers_config
            ]
            logger.info(f"Excel 文件上传并读取成功，共 {len(self.raw_data)} 行数据")
            return True
        except Exception as e:
            self.raw_data = None 
            logger.error(f"Excel 读取失败：{str(e)}")
            raise HTTPException(status_code=400, detail=f"Excel 读取失败：{str(e)}")

    # def run_analysis(self):
    #     """执行所有分析器"""
    #     if self.raw_data is None or not self.analyzers:
    #         raise HTTPException(status_code=400, detail="请先上传Excel文件！")

    #     logger.info("开始执行数据分析...")
        
    #     # 汇总所有分析器的结果
    #     all_analyzed_data = []

    #     # 先执行所有分析器，汇总所有结果
    #     for i, cfg in enumerate(self.analyzers_config):
    #         analyzer = self.analyzers[i]
    #         logger.info(f"开始执行分析器：{analyzer.__class__.__name__}")
    #         success = analyzer.analyze(df=self.raw_data, **cfg["analyze_kwargs"])

    #         if success:
    #             analyzed_df = analyzer.get_analyzed_data()

    #             # 清理特殊的 float 值（NaN, Infinity 等）
    #             cleaned_df = analyzed_df.applymap(self.replace_invalid_floats)

    #             # Convert datetime objects to string
    #             cleaned_data = cleaned_df.to_dict(orient="records")
    #             cleaned_data = convert_datetime_to_string(cleaned_data)  # Apply the conversion

    #             # 汇总所有分析器的结果，并在每个结果中添加分析器名称
    #             all_analyzed_data.append({
    #                 "analyzer_name": analyzer.__class__.__name__,
    #                 "sheet_name": cfg["sheet_name"],
    #                 "status": "success",
    #                 "data": cleaned_data,
    #                 'analyzer':analyzer.__class__.__name__
    #             })
    #             logger.info(f"{analyzer.__class__.__name__} 分析完成，包含 {len(cleaned_df)} 条数据")
    #         else:
    #             logger.error(f"{analyzer.__class__.__name__} 分析失败或无结果")

    #     logger.info("所有分析器执行完毕")
    #     return all_analyzed_data

    def run_analysis(self):
        """执行所有分析器"""
        if self.raw_data is None or not self.analyzers:
            raise HTTPException(status_code=400, detail="请先上传Excel文件！")

        logger.info("开始执行数据分析...")
        
        all_analyzed_data = []
        low_loss_data = []  # ✅ 新增：用于返回前端的亏损<10万元数据

        for i, cfg in enumerate(self.analyzers_config):
            analyzer = self.analyzers[i]
            logger.info(f"开始执行分析器：{analyzer.__class__.__name__}")
            success = analyzer.analyze(df=self.raw_data, **cfg["analyze_kwargs"])

            if success:
                analyzed_df = analyzer.get_analyzed_data()
                cleaned_df = analyzed_df.applymap(self.replace_invalid_floats)
                cleaned_data = cleaned_df.to_dict(orient="records")
                cleaned_data = convert_datetime_to_string(cleaned_data)

                # ✅ 如果是 LossDataAnalyzer，则额外收集亏损<10万元的数据
                if isinstance(analyzer, LossDataAnalyzer):
                    low_loss_df = analyzer.get_low_loss_data()
                    if not low_loss_df.empty:
                        low_loss_cleaned = low_loss_df.applymap(self.replace_invalid_floats)
                        low_loss_cleaned_data = convert_datetime_to_string(
                            low_loss_cleaned.to_dict(orient="records")
                        )
                        low_loss_data = low_loss_cleaned_data  # 保存到变量中

                all_analyzed_data.append({
                    "analyzer_name": analyzer.__class__.__name__,
                    "sheet_name": cfg["sheet_name"],
                    "status": "success",
                    "data": cleaned_data,
                    'analyzer': analyzer.__class__.__name__
                })
                logger.info(f"{analyzer.__class__.__name__} 分析完成，包含 {len(cleaned_df)} 条数据")
            else:
                logger.error(f"{analyzer.__class__.__name__} 分析失败或无结果")

        logger.info("所有分析器执行完毕")

        # ✅ 最终返回时增加一个字段 “low_loss_projects”
        return {
            "all_analyzed_data": all_analyzed_data,
            "low_loss_projects": low_loss_data  # 直接返回给前端
        }

    def classify_projects(self, all_analyzed_data):
        """根据每个项目的异常点数量进行分类"""
        statistics = {
            "one_exception": [],
            "two_exceptions": [],
            "more_than_two_exceptions": [],
            "low_loss_projects":[],
            "all": [],
            "source":[],
        }

        project_stats = {}

        # 遍历所有分析器的结果
        for item in all_analyzed_data:
            cleaned_data = item["data"]
            analyzer_name = item["sheet_name"]
            
            for record in cleaned_data:
                project_name = record.get("项目名称")  # 假设每个项目有一个名称字段

                # 创建一个项目的记录，如果项目不存在，则初始化
                if project_name not in project_stats:
                    project_stats[project_name] = {
                        "exception_count": 0, 
                        "exception_details": [], 
                        "analyzers": set()  # 用于记录分析器名称，使用 set 防止重复
                    }

                # 每个项目的异常点统计：记录该项目被多少个分析器标记为异常
                project_stats[project_name]["exception_count"] += 1
                project_stats[project_name]["exception_details"].append({
                    "field": "项目异常",
                    "exception_type": "由分析器标记"
                })
                
                # 记录哪个分析器认为该项目有异常
                project_stats[project_name]["analyzers"].add(analyzer_name)

        # 根据统计的数量，将每个项目分类
        for project_name, stats in project_stats.items():
            item = {
                "project_name": project_name,
                "exception_count": stats["exception_count"],
                "exception_details": stats["exception_details"],
                "analyzers": list(stats["analyzers"])  # 把 set 转换为 list 以便返回
            }
            statistics["all"].append(item)
            if stats["exception_count"] == 1:
                statistics["one_exception"].append(item)
            elif stats["exception_count"] == 2:
                statistics["two_exceptions"].append(item)
            elif stats["exception_count"] > 2:
                statistics["more_than_two_exceptions"].append(item)
        statistics["source"]=all_analyzed_data
        return statistics

    def replace_invalid_floats(self, value):
        """替换无效的 float 值（NaN, Infinity）"""
        if isinstance(value, float):
            if pd.isna(value) or value == float('inf') or value == float('-inf'):
                return ""  # 或者返回一个替代字符串 "NaN"
        return value

    def save_results_to_excel(self, results: list):
        """将分析结果保存为新的Excel文件"""
        if not results:
            raise HTTPException(status_code=400, detail="没有结果需要保存！")

        logger.info("开始保存分析结果到 Excel 文件...")
        excel_saver = ExcelResultSaver()

        for result in results:
            if result['status'] == 'success' and result['data']:
                try:
                    data_df =result['data']
                    sheet_name = result.get('sheet_name', result['analyzer_name'])

                    # Save data to the sheet
                    excel_saver.save_to_excel(data_df, sheet_base_name=sheet_name)
                    
                except Exception as e:
                    logger.error(f"将 {result['analyzer_name']} 结果保存到 Excel 失败: {e}")
        # Ensure file path is valid
        file_name = "分析报告.xlsx"
        output_dir = "output"
        os.makedirs(output_dir, exist_ok=True)
        file_path = os.path.join(output_dir, file_name)
        # Log file path to check if it's correctly set
        logger.info(f"Saving the result to: {file_path}")
        
        excel_saver.save(file_path)  # Save the final result
        logger.info(f"分析结果已保存到：{file_path}")
        
        return file_path

    def save_results_to_excel_v2(self, results: list):
        """将分析结果保存为新的 Excel 文件 (使用 pandas)"""
        if not results:
            raise HTTPException(status_code=400, detail="没有结果需要保存！")

        output_dir = "output"
        os.makedirs(output_dir, exist_ok=True)
        
        # Set the path for the saved Excel file
        file_name = "分析报告.xlsx"
        file_path = os.path.join(output_dir, file_name)
        
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            for result in results:
                if result["status"] == "success" and result["data"]:
                    sheet_name = result.get("sheet_name", result["analyzer_name"])
                    data_df = pd.DataFrame(result["data"])

                    # Save each result to a separate sheet
                    data_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        logger.info(f"分析结果已保存到：{file_path}")
        # Return the file path
        return file_path
    def get_merged_value(self, sheet, row, col):
        """获取指定行列的单元格值（处理合并单元格）"""
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        return sheet.cell(row=row, column=col).value
    def run_analysis_source(self):
        """执行所有分析器"""
        if self.raw_data is None or not self.analyzers:
            raise HTTPException(status_code=400, detail="请先上传Excel文件！")
        logger.info("开始执行数据分析...")
        # 汇总所有分析器的结果
        all_analyzed_data = []
        # 先执行所有分析器，汇总所有结果
        for i, cfg in enumerate(self.analyzers_config):
            analyzer = self.analyzers[i]
            logger.info(f"开始执行分析器：{analyzer.__class__.__name__}")
            success = analyzer.analyze(df=self.raw_data, **cfg["analyze_kwargs"])
            if success:
                analyzed_df = analyzer.get_analyzed_data()
                # 清理特殊的 float 值（NaN, Infinity 等）
                cleaned_df = analyzed_df.applymap(self.replace_invalid_floats)

                # Convert datetime objects to string
                cleaned_data = cleaned_df.to_dict(orient="records")
                cleaned_data = convert_datetime_to_string(cleaned_data)  # Apply the conversion

                # 汇总所有分析器的结果，并在每个结果中添加分析器名称
                all_analyzed_data.append({
                    "analyzer_name": analyzer.__class__.__name__,
                    "sheet_name": cfg["sheet_name"],
                    "status": "success",
                    "data": cleaned_data,
                    'analyzer':analyzer.__class__.__name__
                })
                logger.info(f"{analyzer.__class__.__name__} 分析完成，包含 {len(cleaned_df)} 条数据")
            else:
                logger.error(f"{analyzer.__class__.__name__} 分析失败或无结果")

        logger.info("所有分析器执行完毕")
        return all_analyzed_data

analysis_api = AnalysisAPI()    

def convert_all_non_json_compliant_to_string(obj):
    """Recursively convert all non-JSON-compliant types to string."""
    if isinstance(obj, (float, int)):  # 处理 float 和 int 类型
        if pd.isna(obj) or obj == float('inf') or obj == float('-inf'):
            return str(obj)  # 将 NaN 和 Infinity 转换为字符串
        return str(obj)  # 转换所有数字为字符串
    elif isinstance(obj, datetime):  # 处理 datetime 类型
        return obj.strftime('%Y-%m-%d %H:%M:%S')  # 转换为字符串
    elif isinstance(obj, dict):  # 如果是字典，递归处理其值
        return {key: convert_all_non_json_compliant_to_string(value) for key, value in obj.items()}
    elif isinstance(obj, list):  # 如果是列表，递归处理每个元素
        return [convert_all_non_json_compliant_to_string(item) for item in obj]
    return str(obj)  # 处理其他类型，直接转换为字符串

@app.post("/upload_and_analyze_json/", tags=["一站式API"])
async def upload_and_analyze_json(file: UploadFile = File(..., description="要分析的项目数据 Excel 文件")):
    """
    【一站式】上传 Excel 文件，立即执行所有分析，并返回 JSON 格式的结果。
    """
    try:
        analysis_api.upload_excel(file)
        results = analysis_api.run_analysis()
        # 这里的 run_analysis() 现在返回 dict，包括：
        # { "all_analyzed_data": [...], "low_loss_projects": [...] }

        all_analyzed_data = results["all_analyzed_data"]
        low_loss_projects = results.get("low_loss_projects", [])

        # 对主要分析数据执行分类统计
        classified_results = analysis_api.classify_projects(all_analyzed_data)

        # 把低额亏损项目附加进最终返回结果
        classified_results["low_loss_projects"] = low_loss_projects

        # 转换为可序列化结构
        classified_results = convert_all_non_json_compliant_to_string(classified_results)
        classified_results = jsonable_encoder(classified_results)
        return JSONResponse(content=classified_results)
    except HTTPException as e:
        logger.error(f"HTTP 错误：{e.detail}")
        raise e
    except Exception as e:
        logger.error(f"发生未知错误：{str(e)}")
        if "Out of range float values" in str(e):
            raise HTTPException(status_code=500, detail="分析结果包含非标准的浮动数 (NaN/Inf)，请检查数据清洗。")
        raise HTTPException(status_code=500, detail=f"处理请求时发生未知错误: {str(e)}")

@app.post("/upload_and_download_excel/", tags=["一站式API"])
async def upload_and_download_excel(file: UploadFile = File(..., description="要分析的项目数据 Excel 文件")):
    """
    【一站式】上传 Excel 文件，执行分析，并将结果保存为 Excel 文件并直接返回下载。
    """
    try:
        logger.info("开始上传并分析 Excel 文件...")
        
        # 上传并解析Excel文件
        analysis_api.upload_excel(file)

        # 执行分析并获取所有分析的结果
        all_analyzed_data = analysis_api.run_analysis()

        # Use the new save method to save results to Excel
        file_path = analysis_api.save_results_to_excel_v2(all_analyzed_data)

        # Ensure that the file exists before sending the response
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="生成的Excel文件未找到")

        # Return the file as a response to the client
        return FileResponse(
            file_path, 
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
            filename="分析报告.xlsx"
        )

    except HTTPException as e:
        logger.error(f"HTTP 错误：{e.detail}")
        raise e
    except Exception as e:
        logger.error(f"发生未知错误：{str(e)}")
        raise HTTPException(status_code=500, detail=f"处理请求时发生未知错误: {str(e)}")

from fastapi import Body

@app.post("/download_excel/", tags=["一站式API"])
async def upload_and_download_excel(all_analyzed_data: list = Body(...)):
    """
    【一站式】上传 Excel 文件，执行分析，并将结果保存为 Excel 文件并直接返回下载。
    """
    try:
        logger.info("开始上传并分析 Excel 文件...")

        file_path = analysis_api.save_results_to_excel_v2(all_analyzed_data)

        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="生成的Excel文件未找到")

        return FileResponse(
            file_path,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename="分析报告.xlsx"
        )

    except HTTPException as e:
        logger.error(f"HTTP 错误：{e.detail}")
        raise e
    except Exception as e:
        logger.error(f"发生未知错误：{str(e)}")
        raise HTTPException(status_code=500, detail=f"处理请求时发生未知错误: {str(e)}")


# 运行FastAPI服务器
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8004)
